from __future__ import annotations

from datetime import date

from httpx import AsyncClient
from sqlalchemy import select

from app.db import SessionLocal
from app.models.enums import Register
from app.models.master import Vehicle, WorkType
from tests.conftest import auth_headers

FROM = date(2026, 8, 1)
TO = date(2026, 8, 10)
BUS = "MH40LY1894"
OTHER = "MH40LY1895"


def window() -> dict[str, str]:
    return {"from": FROM.isoformat(), "to": TO.isoformat()}


def row(body: dict, reg: str) -> dict:
    return next(r for r in body["rows"] if r["registration_no"] == reg)


def cell(body: dict, reg: str, day: int) -> dict:
    """The block for one bus on one day of August."""
    index = body["dates"].index(date(2026, 8, day).isoformat())
    return row(body, reg)["cells"][index]


async def chart(client: AsyncClient, headers: dict, kind: str) -> dict:
    resp = await client.get(
        f"/sites/MBMT/reports/control-charts/{kind}",
        params=window(),
        headers=headers,
    )
    assert resp.status_code == 200, resp.text
    return resp.json()


async def _inspection_types() -> dict[str, int]:
    async with SessionLocal() as session:
        ids: dict[str, int] = {}
        for code, name in (
            ("D.I", "Daily inspection"),
            ("10 DAYS SERVICE", "10 day inspection"),
            ("P.M", "Preventive maintenance docking"),
        ):
            work_type = WorkType(code=code, name=name, is_inspection=True)
            session.add(work_type)
            await session.flush()
            ids[code] = work_type.id
        session.add(WorkType(code="B.D", name="Breakdown", register=Register.breakdown))
        await session.commit()
        return ids


async def _vehicle_id(reg: str = BUS) -> str:
    async with SessionLocal() as session:
        return (
            await session.scalar(
                select(Vehicle).where(Vehicle.registration_no == reg)
            )
        ).id


async def _inspect(
    client: AsyncClient,
    headers: dict,
    work_type_id: int,
    day: int,
    *,
    reg: str = BUS,
    results: list[dict] | None = None,
) -> None:
    resp = await client.post(
        "/sites/MBMT/inspections",
        json={
            "vehicle_id": await _vehicle_id(reg),
            "work_type_id": work_type_id,
            "inspected_on": date(2026, 8, day).isoformat(),
            "results": results or [],
        },
        headers=headers,
    )
    assert resp.status_code == 201, resp.text


async def _entry(
    client: AsyncClient,
    headers: dict,
    register: str,
    day: int,
    data: dict,
    *,
    reg: str = BUS,
) -> None:
    resp = await client.post(
        "/entries",
        json={
            "register": register,
            "site": "MBMT",
            "date": date(2026, 8, day).isoformat(),
            "data": {"bus_no": reg, **data},
        },
        headers=headers,
    )
    assert resp.status_code == 201, resp.text


# --- the shape every chart shares --------------------------------------------


async def test_a_chart_is_the_whole_fleet_by_every_day_in_the_window(
    client: AsyncClient,
) -> None:
    """A grid with buses missing is the one thing a control chart cannot be —
    the empty row is the finding."""
    h = await auth_headers(client)
    body = await chart(client, h, "coolantTopping")

    assert len(body["dates"]) == 10
    # Two active buses at the fixture site; the retired one is not fleet.
    assert [r["registration_no"] for r in body["rows"]] == [BUS, OTHER]
    assert all(len(r["cells"]) == 10 for r in body["rows"])
    assert body["filled"] == 0


async def test_a_window_wider_than_a_month_is_refused(client: AsyncClient) -> None:
    h = await auth_headers(client)
    resp = await client.get(
        "/sites/MBMT/reports/control-charts/coolantTopping",
        params={"from": "2026-08-01", "to": "2026-09-30"},
        headers=h,
    )
    assert resp.status_code == 400
    assert "31 days" in resp.json()["error"]["message"]


async def test_a_backwards_window_is_refused(client: AsyncClient) -> None:
    h = await auth_headers(client)
    resp = await client.get(
        "/sites/MBMT/reports/control-charts/coolantTopping",
        params={"from": "2026-08-10", "to": "2026-08-01"},
        headers=h,
    )
    assert resp.status_code == 400


async def test_a_site_you_cannot_reach_is_refused(client: AsyncClient) -> None:
    h = await auth_headers(client, "TV4105")
    resp = await client.get(
        "/sites/UMT/reports/control-charts/coolantTopping",
        params=window(),
        headers=h,
    )
    assert resp.status_code == 403


# --- coolant topping ---------------------------------------------------------


async def test_coolant_topping_shows_litres_and_shades_pm_days(
    client: AsyncClient,
) -> None:
    h = await auth_headers(client)
    types = await _inspection_types()
    await _entry(
        client, h, "coolant", 3, {"bcs_litres": 1.5, "tcs_litres": 0.5}
    )
    await _entry(client, h, "coolant", 7, {"bcs_litres": 2})
    await _inspect(client, h, types["10 DAYS SERVICE"], 7)

    body = await chart(client, h, "coolantTopping")
    assert cell(body, BUS, 3) == {"value": "2", "mark": "plain", "title": ""}
    # Topped up on the day it was serviced — the depot shades that block so the
    # two are not read as unrelated.
    assert cell(body, BUS, 7) == {"value": "2", "mark": "pm", "title": ""}
    assert cell(body, BUS, 4) == {"value": "", "mark": "plain", "title": ""}
    assert cell(body, OTHER, 3) == {"value": "", "mark": "plain", "title": ""}


async def test_coolant_topping_adds_up_two_toppings_in_one_day(
    client: AsyncClient,
) -> None:
    h = await auth_headers(client)
    await _entry(client, h, "coolant", 5, {"bcs_litres": 1.5})
    await _entry(client, h, "coolant", 5, {"tcs_litres": 2.5})

    body = await chart(client, h, "coolantTopping")
    assert cell(body, BUS, 5)["value"] == "4"


async def test_a_pm_day_with_no_topping_is_still_shaded(
    client: AsyncClient,
) -> None:
    """The shading says "serviced", not "topped up"; losing it would hide the
    days a service needed no coolant at all."""
    h = await auth_headers(client)
    types = await _inspection_types()
    await _inspect(client, h, types["D.I"], 4)

    body = await chart(client, h, "coolantTopping")
    assert cell(body, BUS, 4) == {"value": "", "mark": "pm", "title": ""}


# --- D.I and 10-day service --------------------------------------------------


async def test_di_inspection_ticks_only_di_days(client: AsyncClient) -> None:
    h = await auth_headers(client)
    types = await _inspection_types()
    await _inspect(client, h, types["D.I"], 2)
    await _inspect(client, h, types["10 DAYS SERVICE"], 6)

    body = await chart(client, h, "diInspection")
    assert cell(body, BUS, 2) == {"value": "✓", "mark": "plain", "title": ""}
    assert cell(body, BUS, 6) == {"value": "", "mark": "plain", "title": ""}


async def test_ten_day_service_ticks_only_ten_day_days(client: AsyncClient) -> None:
    h = await auth_headers(client)
    types = await _inspection_types()
    await _inspect(client, h, types["D.I"], 2)
    await _inspect(client, h, types["10 DAYS SERVICE"], 6)

    body = await chart(client, h, "tenDayService")
    assert cell(body, BUS, 2) == {"value": "", "mark": "plain", "title": ""}
    assert cell(body, BUS, 6) == {"value": "✓", "mark": "plain", "title": ""}


async def test_di_and_ten_day_both_tick_when_a_bus_has_both_the_same_day(
    client: AsyncClient,
) -> None:
    """Each chart reads its own work-type code — a D.I and a ten-day service
    on the same day must not fight over which one shows."""
    h = await auth_headers(client)
    types = await _inspection_types()
    await _inspect(client, h, types["D.I"], 5)
    await _inspect(client, h, types["10 DAYS SERVICE"], 5)

    assert cell(await chart(client, h, "diInspection"), BUS, 5)["value"] == "✓"
    assert cell(await chart(client, h, "tenDayService"), BUS, 5)["value"] == "✓"


# --- driver complaints and breakdowns -----------------------------------------


async def test_driver_complaints_chart_carries_the_complaint_text(
    client: AsyncClient,
) -> None:
    h = await auth_headers(client)
    await _entry(client, h, "driver_complaint", 2, {"complaint": "AC weak"})

    body = await chart(client, h, "driverComplaints")
    assert cell(body, BUS, 2) == {"value": "C", "mark": "plain", "title": "AC weak"}


async def test_driver_complaints_chart_counts_and_joins_multiple_same_day(
    client: AsyncClient,
) -> None:
    h = await auth_headers(client)
    await _entry(client, h, "driver_complaint", 2, {"complaint": "AC weak"})
    await _entry(client, h, "driver_complaint", 2, {"complaint": "Door slow"})

    body = await chart(client, h, "driverComplaints")
    cell_ = cell(body, BUS, 2)
    assert cell_["value"] == "2"
    assert cell_["title"] == "AC weak; Door slow"


async def test_breakdowns_chart_is_its_own_grid_not_folded_into_complaints(
    client: AsyncClient,
) -> None:
    h = await auth_headers(client)
    await _entry(client, h, "driver_complaint", 8, {"complaint": "Noise"})
    await _entry(client, h, "breakdown", 8, {"complaint": "No traction"})

    breakdowns = await chart(client, h, "breakdowns")
    assert cell(breakdowns, BUS, 8) == {
        "value": "BD",
        "mark": "breakdown",
        "title": "No traction",
    }
    # A breakdown on the same day no longer erases the complaint's own chart.
    complaints = await chart(client, h, "driverComplaints")
    assert cell(complaints, BUS, 8) == {
        "value": "C",
        "mark": "plain",
        "title": "Noise",
    }


# --- the two charts a checklist line answers ---------------------------------


async def _checklist_with_chart_lines(
    client: AsyncClient, headers: dict, work_type_id: int
) -> dict[str, str]:
    """A D.I checklist where two lines are nominated to feed the charts."""
    resp = await client.put(
        f"/sites/MBMT/checklists/{work_type_id}",
        json={
            "name": "Daily inspection",
            "items": [
                {"section": "Safety", "label": "Horn working"},
                {
                    "section": "Tyres",
                    "label": "Tyre pressure checked with gauge",
                    "chart_key": "tyre_pressure",
                },
                {
                    "section": "Cleaning",
                    "label": "Bus washed",
                    "chart_key": "washing",
                },
            ],
        },
        headers=headers,
    )
    assert resp.status_code == 200, resp.text
    return {i["label"]: i["id"] for i in resp.json()["items"]}


async def test_a_nominated_checklist_line_ticks_its_chart(
    client: AsyncClient,
) -> None:
    """Neither chart needs its own data entry — the line already on the daily
    inspection is the record."""
    h = await auth_headers(client)
    types = await _inspection_types()
    items = await _checklist_with_chart_lines(client, h, types["D.I"])

    await _inspect(
        client,
        h,
        types["D.I"],
        3,
        results=[
            {"item_id": items["Horn working"], "result": "ok"},
            {"item_id": items["Tyre pressure checked with gauge"], "result": "ok"},
            {"item_id": items["Bus washed"], "result": "ok"},
        ],
    )

    tyres = await chart(client, h, "tyrePressure")
    assert cell(tyres, BUS, 3) == {"value": "✓", "mark": "plain", "title": ""}
    assert cell(tyres, BUS, 4) == {"value": "", "mark": "plain", "title": ""}

    washing = await chart(client, h, "busWashing")
    assert cell(washing, BUS, 3)["value"] == "✓"


async def test_a_line_answered_not_applicable_is_not_a_tick(
    client: AsyncClient,
) -> None:
    """"Not applicable" is how a skipped check is recorded honestly; a tick
    would say the opposite of what was entered."""
    h = await auth_headers(client)
    types = await _inspection_types()
    items = await _checklist_with_chart_lines(client, h, types["D.I"])

    await _inspect(
        client,
        h,
        types["D.I"],
        3,
        results=[
            {"item_id": items["Horn working"], "result": "ok"},
            {"item_id": items["Tyre pressure checked with gauge"], "result": "na"},
            {"item_id": items["Bus washed"], "result": "not_ok"},
        ],
    )

    assert cell(await chart(client, h, "tyrePressure"), BUS, 3)["value"] == ""
    # Washed badly is still washed — only "not applicable" means it did not
    # happen.
    assert cell(await chart(client, h, "busWashing"), BUS, 3)["value"] == "✓"


async def test_an_inspection_with_no_nominated_line_leaves_the_chart_empty(
    client: AsyncClient,
) -> None:
    """Until the depot marks the line, the chart says nothing rather than
    guessing from the wording."""
    h = await auth_headers(client)
    types = await _inspection_types()
    resp = await client.put(
        f"/sites/MBMT/checklists/{types['D.I']}",
        json={
            "name": "Daily inspection",
            "items": [{"section": "Tyres", "label": "Tyre pressure"}],
        },
        headers=h,
    )
    item_id = resp.json()["items"][0]["id"]
    await _inspect(
        client,
        h,
        types["D.I"],
        3,
        results=[{"item_id": item_id, "result": "ok"}],
    )

    body = await chart(client, h, "tyrePressure")
    assert body["filled"] == 0


# --- what the system cannot answer -------------------------------------------


async def test_the_energy_chart_says_why_it_is_empty(client: AsyncClient) -> None:
    """A blank grid and an unanswerable one look identical, so the chart that
    has no feed behind it says so."""
    h = await auth_headers(client)
    body = await chart(client, h, "energy")
    assert body["available"] is False
    assert "energy" in body["unavailable_reason"].lower()
    assert body["filled"] == 0


async def test_the_chart_list_marks_which_ones_have_data_behind_them(
    client: AsyncClient,
) -> None:
    h = await auth_headers(client)
    resp = await client.get("/reports/control-charts", headers=h)
    assert resp.status_code == 200, resp.text
    charts = {c["kind"]: c for c in resp.json()}
    assert len(charts) == 8
    assert charts["energy"]["available"] is False
    assert [k for k, c in charts.items() if not c["available"]] == ["energy"]
    assert charts["diInspection"]["title"] == "D.I inspection"


# --- the export --------------------------------------------------------------


async def test_the_export_carries_the_colour_a_csv_cannot_hold(
    client: AsyncClient,
) -> None:
    h = await auth_headers(client)
    await _entry(client, h, "breakdown", 4, {"complaint": "Air leak"})

    resp = await client.get(
        "/sites/MBMT/reports/control-charts/breakdowns/export",
        params=window(),
        headers=h,
    )
    assert resp.status_code == 200, resp.text
    assert "text/csv" in resp.headers["content-type"]
    body = resp.text
    assert "Breakdowns" in body
    assert "Air leak (BD)" in body
    assert BUS in body and OTHER in body


async def test_breakdowns_and_driver_complaints_export_as_excel(
    client: AsyncClient,
) -> None:
    """The only two charts a block's full text is worth reading out of a
    spreadsheet for."""
    import io

    from openpyxl import load_workbook

    h = await auth_headers(client)
    await _entry(client, h, "breakdown", 4, {"complaint": "Air leak"})
    await _entry(client, h, "driver_complaint", 4, {"complaint": "AC weak"})

    for kind, text in (("breakdowns", "Air leak"), ("driverComplaints", "AC weak")):
        resp = await client.get(
            f"/sites/MBMT/reports/control-charts/{kind}/export",
            params={**window(), "format": "xlsx"},
            headers=h,
        )
        assert resp.status_code == 200, resp.text
        assert resp.headers["content-type"] == (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        sheet = load_workbook(io.BytesIO(resp.content)).active
        rows = list(sheet.iter_rows(values_only=True))
        header = next(r for r in rows if r[0] == "Bus No")
        day_col = header.index("2026-08-04")
        bus_row = next(r for r in rows if r[0] == BUS)
        assert bus_row[day_col] == text


async def test_other_charts_ignore_the_excel_format(client: AsyncClient) -> None:
    """Only breakdowns and driver complaints have an Excel writer — anything
    else asking for xlsx still gets the CSV, same as any unhandled format."""
    h = await auth_headers(client)
    resp = await client.get(
        "/sites/MBMT/reports/control-charts/coolantTopping/export",
        params={**window(), "format": "xlsx"},
        headers=h,
    )
    assert resp.status_code == 200, resp.text
    assert "text/csv" in resp.headers["content-type"]
