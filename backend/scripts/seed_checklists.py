"""Provision a site's inspection checklists from its own sheets.

The checks are the depot's, not this file's: they are read off the sheet's
header row, in the order the sheet has them, so what a mechanic sees on screen
is the row they used to tick on paper.

MBMT's sheet is three sheets — 9M, 12M AC and 12M non-AC — sharing eleven
checks and differing in the rest. Each becomes a template scoped to a variant,
and the buses each sheet lists are marked as taking it. That mapping is the
sheet's own: which bus is air-conditioned is not otherwise written down
anywhere in the system.

The two sheets are laid out differently, and both are read as they are rather
than reformatted: the daily sheet runs its checks across the columns with a bus
per row, and the ten-day sheet runs them down the rows with a check location
beside each. That location becomes the section a check sits under on screen.

    python -m scripts.seed_checklists [SITE_CODE] [DATA_DIR]

Idempotent: re-running replaces the lines, keeps any a result already points
at, and re-marks the fleet.
"""
from __future__ import annotations

import asyncio
import sys
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import func, select

from app.db import SessionLocal
from app.models.master import Vehicle, WorkType
from app.services import checklists
from app.services.imports import normalize_registration_no

DEFAULT_SITE = "MBMT"
DEFAULT_DIR = "/srv/data/MBMT/August"

DI_SHEET = "D.I. SHEET.xlsx"

#: Worksheet -> the variant it describes. A sheet that is only a continuation
#: of another (the 9M fleet runs onto a second page) names the same variant and
#: contributes its buses.
VARIANTS: dict[str, str] = {
    "9M D.i Sheet 1": "9M",
    "9M D.i Sheet 2": "9M",
    "12M AC  D.I SHEET": "12M AC",
    "12M  Non-AC  D.I SHEET": "12M Non-AC",
}

#: The ten-day sheets, and the variants each one covers. One sheet serves both
#: 12M fleets — the depot checks an air-conditioned 12M and a non-AC one the
#: same way at ten days, unlike the daily.
TEN_DAY_SHEETS: list[tuple[str, tuple[str, ...]]] = [
    ("10 Days 9M (new).xlsx", ("9M",)),
    ("10 Days 12M AC & NAC FORMAT (2).xlsx", ("12M AC", "12M Non-AC")),
]

#: Where the ten-day sheets keep their columns.
TEN_DAY_DESCRIPTION = 1
TEN_DAY_LOCATION = 5

#: Columns that are not checks: what identifies the bus, and the free-text
#: tail. Matched on the leading word too, because the corner cell carries the
#: sheet's own furniture — "Checks", but also "Checks Date:" on one variant.
NOT_A_CHECK = {"sr no", "sr.no", "bus no.", "bus no", "km", "checks", "remarks"}
NOT_A_CHECK_PREFIX = ("checks", "date", "depot", "name of", "technician")

#: The row the checks are written across on every one of these sheets.
HEADER_ROW = 3

#: Registration numbers on the D.I sheet are the short form the depot speaks —
#: "5736" for MH04LQ5736 — so a bus is matched on the tail of its number.
MIN_SUFFIX = 3


def _clean(value: object) -> str:
    return " ".join(str(value or "").split())


def _is_check(label: str) -> bool:
    if not label:
        return False
    lowered = label.lower()
    if lowered in NOT_A_CHECK:
        return False
    return not lowered.startswith(NOT_A_CHECK_PREFIX)


def read_sheet(path: Path) -> dict[str, tuple[list[str], set[str]]]:
    """variant -> (its checks in sheet order, the bus numbers that take it)."""
    workbook = load_workbook(path, data_only=True, read_only=True)
    found: dict[str, tuple[list[str], set[str]]] = {}

    for name in workbook.sheetnames:
        variant = VARIANTS.get(name)
        if variant is None:
            continue
        worksheet = workbook[name]
        rows = list(worksheet.iter_rows(values_only=True))
        if len(rows) < HEADER_ROW:
            continue

        header = rows[HEADER_ROW - 1]
        checks = [
            _clean(cell) for cell in header if _is_check(_clean(cell))
        ]

        # The bus column is the second on every variant of the sheet.
        buses = set()
        for row in rows[HEADER_ROW:]:
            if len(row) < 2:
                continue
            raw = _clean(row[1])
            if raw and raw.replace(".", "").isdigit():
                buses.add(raw.split(".")[0])

        checks_so_far, buses_so_far = found.get(variant, ([], set()))
        found[variant] = (checks_so_far or checks, buses_so_far | buses)

    return found


def read_ten_day(path: Path) -> list[tuple[str, str]]:
    """(section, check) down the rows, in the sheet's order.

    The sheet's "Check Location" — Inside Bus, Under Carriage, Re-verify —
    becomes the section a check sits under, so the form groups the way the
    mechanic walks the bus.
    """
    workbook = load_workbook(path, data_only=True, read_only=True)
    rows = list(workbook.worksheets[0].iter_rows(values_only=True))

    items: list[tuple[str, str]] = []
    for row in rows:
        if len(row) <= TEN_DAY_DESCRIPTION:
            continue
        description = _clean(row[TEN_DAY_DESCRIPTION])
        location = (
            _clean(row[TEN_DAY_LOCATION]) if len(row) > TEN_DAY_LOCATION else ""
        )
        # The header row names the column rather than a job, and everything
        # above it is the sheet's letterhead.
        if not description or description.lower() == "job description":
            continue
        if not location:
            continue
        items.append((location, description))
    return items


def match_fleet(fleet: list[Vehicle], numbers: set[str]) -> list[Vehicle]:
    """The buses those short numbers name.

    The sheet writes "5736" where the fleet holds MH04LQ5736, so the match is
    on the tail. Short enough numbers would match several buses and are left
    alone rather than guessed at.
    """
    matched: list[Vehicle] = []
    for number in numbers:
        if len(number) < MIN_SUFFIX:
            continue
        hits = [
            v
            for v in fleet
            if normalize_registration_no(v.registration_no).endswith(number)
        ]
        if len(hits) == 1:
            matched.append(hits[0])
    return matched


async def _work_type(session, code: str) -> WorkType:
    found = await session.scalar(
        select(WorkType).where(func.upper(WorkType.code) == code.upper())
    )
    if found is None:
        raise SystemExit(f"No {code} work type — run scripts.seed first")
    return found


async def seed(site_code: str, folder: Path) -> None:
    sheets = read_sheet(folder / DI_SHEET)
    if not sheets:
        raise SystemExit(f"No recognised D.I sheets in {folder / DI_SHEET}")

    async with SessionLocal() as session:
        work_type = await _work_type(session, "D.I")

        fleet = list(
            (
                await session.scalars(
                    select(Vehicle).where(Vehicle.site_code == site_code)
                )
            )
            .unique()
            .all()
        )

        for variant, (checks, numbers) in sorted(sheets.items()):
            template = await checklists.ensure_template(
                session, site_code, work_type, variant=variant
            )
            template.name = f"Daily inspection — {variant}"
            await checklists.replace_items(
                session,
                template,
                [
                    checklists.ChecklistLine(section="", label=check)
                    for check in checks
                ],
            )

            buses = match_fleet(fleet, numbers)
            for bus in buses:
                bus.checklist_variant = variant
            print(
                f"  {variant:12s} {len(checks):2d} checks, {len(buses):2d} buses"
            )

        # --- the ten-day service ---
        ten_day = await _work_type(session, "10 DAYS SERVICE")
        for file_name, variants in TEN_DAY_SHEETS:
            path = folder / file_name
            if not path.exists():
                print(f"  (no {file_name}, skipping)")
                continue
            items = read_ten_day(path)
            for variant in variants:
                template = await checklists.ensure_template(
                    session, site_code, ten_day, variant=variant
                )
                template.name = f"10 day inspection — {variant}"
                await checklists.replace_items(
                    session,
                    template,
                    [
                        checklists.ChecklistLine(section=section, label=label)
                        for section, label in items
                    ],
                )
                print(f"  10-day {variant:12s} {len(items):2d} checks")

        unassigned = [v for v in fleet if not v.checklist_variant]
        if unassigned:
            print(
                f"  {len(unassigned)} bus(es) name no variant and will take the "
                "site's unscoped checklist if it has one: "
                + ", ".join(v.registration_no for v in unassigned[:6])
            )
        await session.commit()
    print("Inspection checklists ready.")


if __name__ == "__main__":
    site = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_SITE
    folder = Path(sys.argv[2] if len(sys.argv) > 2 else DEFAULT_DIR)
    asyncio.run(seed(site, folder))
