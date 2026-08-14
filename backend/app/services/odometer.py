from __future__ import annotations

import logging
from dataclasses import dataclass
from datetime import UTC, datetime
from datetime import date as date_t
from pathlib import Path

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.errors import ValidationError
from app.models.master import OdometerReading, Vehicle
from app.models.site_config import SiteConfig
from app.schemas.site import OdometerReadingOut, OdometerSyncOut
from app.services.site_config import get_or_create

logger = logging.getLogger("enm.odometer")


class TelematicsProvider:
    """Where odometer readings come from.

    No provider is configured out of the box, and that is a supported state:
    sites without a feed rely on manual readings and the odometer import, and
    the rest of the schedule still works. A pull that finds nothing new must
    succeed with an empty list, never 500.
    """

    def __init__(self, source: str) -> None:
        self.source = source

    @property
    def is_configured(self) -> bool:
        return False

    async def fetch(
        self, _registration_nos: list[str]
    ) -> dict[str, tuple[int, datetime]]:
        return {}


def provider_for(config: SiteConfig) -> TelematicsProvider:
    """The feed this site reads from.

    `file:/path/to/VehicleStatus.xlsx` reads the depot's own export; anything
    else is the unconfigured default, which succeeds with nothing.
    """
    source = config.odometer_sync_source or ""
    if source.startswith(FILE_SOURCE_PREFIX):
        return VehicleStatusFileProvider(source)
    return TelematicsProvider(source)


def record_reading(
    session: AsyncSession,
    vehicle: Vehicle,
    *,
    odometer_km: int,
    recorded_at: datetime,
    source: str,
) -> None:
    """Cache the latest on the vehicle and append to the history."""
    vehicle.odometer_km = odometer_km
    vehicle.odometer_updated_at = recorded_at
    session.add(
        OdometerReading(
            vehicle_id=vehicle.id,
            odometer_km=odometer_km,
            recorded_at=recorded_at,
            source=source,
        )
    )


async def set_manual(
    session: AsyncSession, vehicle: Vehicle, odometer_km: int
) -> Vehicle:
    """An odometer never moves backwards — a lower manual reading is rejected."""
    if vehicle.odometer_updated_at is not None and odometer_km < vehicle.odometer_km:
        raise ValidationError(
            f"Odometers do not run backwards — {vehicle.registration_no} is "
            f"already at {vehicle.odometer_km:,} km",
            {"odometer_km": "lower than the reading on record"},
        )
    record_reading(
        session,
        vehicle,
        odometer_km=odometer_km,
        recorded_at=datetime.now(UTC),
        source="manual",
    )
    await session.flush()
    return vehicle


async def record_service(
    session: AsyncSession,
    vehicle: Vehicle,
    *,
    plan_code: str,
    odometer_km: int,
    serviced_on: date_t,
) -> Vehicle:
    """Close out a service and re-anchor the next one."""
    if vehicle.odometer_updated_at is not None and odometer_km < vehicle.odometer_km:
        raise ValidationError(
            f"Odometers do not run backwards — {vehicle.registration_no} is "
            f"already at {vehicle.odometer_km:,} km",
            {"odometer_km": "lower than the reading on record"},
        )
    record_reading(
        session,
        vehicle,
        odometer_km=odometer_km,
        recorded_at=datetime.now(UTC),
        source="service",
    )
    vehicle.last_service_km = odometer_km
    vehicle.last_service_on = serviced_on
    vehicle.last_service_code = plan_code.strip().upper()
    await session.flush()
    return vehicle


async def sync_site(session: AsyncSession, site_code: str) -> OdometerSyncOut:
    """Pull the latest readings for one site's fleet.

    Idempotent: the client polls on the same interval as the server job, so a
    pull that finds nothing new reports `skipped`, not an error.
    """
    config = await get_or_create(session, site_code)
    now = datetime.now(UTC)

    vehicles = list(
        (
            await session.scalars(
                select(Vehicle).where(
                    Vehicle.site_code == site_code, Vehicle.is_active.is_(True)
                )
            )
        ).all()
    )

    provider = provider_for(config)
    fetched = (
        await provider.fetch([v.registration_no for v in vehicles])
        if provider.is_configured
        else {}
    )

    updated: list[OdometerReadingOut] = []
    skipped = 0
    for vehicle in vehicles:
        reading = fetched.get(vehicle.registration_no)
        # Nothing reported, or not newer than what we hold: not an error.
        if reading is None or reading[0] <= vehicle.odometer_km:
            skipped += 1
            continue
        odometer_km, recorded_at = reading
        record_reading(
            session,
            vehicle,
            odometer_km=odometer_km,
            recorded_at=recorded_at,
            source=config.odometer_sync_source,
        )
        updated.append(
            OdometerReadingOut(
                vehicle_id=vehicle.id,
                registration_no=vehicle.registration_no,
                odometer_km=odometer_km,
                recorded_at=recorded_at,
            )
        )

    config.odometer_last_synced_at = now
    await session.flush()
    return OdometerSyncOut(readings=updated, synced_at=now, skipped=skipped)


async def scan_sites_due_for_sync() -> None:
    """Scheduled job: pull for every site whose interval has elapsed.

    Runs alongside the breakdown-SLA scan. The client polls too, which is why
    the pull has to be idempotent.
    """
    from app.db import SessionLocal

    async with SessionLocal() as session:
        configs = list(
            (
                await session.scalars(
                    select(SiteConfig).where(SiteConfig.odometer_sync_enabled.is_(True))
                )
            ).all()
        )
        now = datetime.now(UTC)
        for config in configs:
            last = config.odometer_last_synced_at
            if last is not None:
                elapsed_minutes = (now - last).total_seconds() / 60
                if elapsed_minutes < config.odometer_sync_minutes:
                    continue
            try:
                result = await sync_site(session, config.site_code)
                await session.commit()
            except Exception:  # noqa: BLE001 - one bad site must not stop the rest
                await session.rollback()
                logger.exception("Odometer sync failed for %s", config.site_code)
                continue
            if result.readings:
                logger.info(
                    "Odometer sync %s: %d updated, %d skipped",
                    config.site_code,
                    len(result.readings),
                    result.skipped,
                )


# --- a feed that reads the depot's own vehicle-status export ------------------

#: Where the columns sit on MBMT's VehicleStatus export. The sheet opens with a
#: subreport error and a title, so the headings are on row 4 and the fleet
#: starts on row 5.
STATUS_HEADER_ROW = 4
STATUS_REGISTRATION = 2
STATUS_VEHICLE_TYPE = 3
STATUS_DATE = 4
STATUS_STATUS = 5
STATUS_END_ODOMETER = 8

#: `odometer_sync_source` values that mean "read the export at this path".
FILE_SOURCE_PREFIX = "file:"


@dataclass(frozen=True, slots=True)
class VehicleStatusRow:
    """One line of the export: what a bus is and where its odometer stands."""

    registration_no: str
    vehicle_type: str
    odometer_km: int | None
    status: str
    recorded_at: datetime | None


def read_vehicle_status(path: Path) -> list[VehicleStatusRow]:
    """Parse a VehicleStatus export.

    Readings arrive as floats — 104997.3, 78635.203125 — because they come off
    a telematics feed rather than a dial. They are floored rather than rounded:
    a bus has not done the next kilometre until it has done it.

    A zero reading means the feed has nothing for that bus, not that the bus
    has never moved, so it comes back as None and the caller leaves the vehicle
    alone.
    """
    from openpyxl import load_workbook  # noqa: PLC0415 — optional at import time

    workbook = load_workbook(path, data_only=True, read_only=True)
    worksheet = workbook.worksheets[0]

    rows: list[VehicleStatusRow] = []
    for row in worksheet.iter_rows(min_row=STATUS_HEADER_ROW + 1, values_only=True):
        if len(row) <= STATUS_END_ODOMETER:
            continue
        registration = " ".join(str(row[STATUS_REGISTRATION] or "").split())
        if not registration:
            continue

        raw = row[STATUS_END_ODOMETER]
        km: int | None = None
        if isinstance(raw, int | float) and raw > 0:
            km = int(raw)

        stamp = row[STATUS_DATE]
        rows.append(
            VehicleStatusRow(
                registration_no=registration.upper(),
                vehicle_type=" ".join(str(row[STATUS_VEHICLE_TYPE] or "").split()),
                odometer_km=km,
                status=" ".join(str(row[STATUS_STATUS] or "").split()),
                recorded_at=stamp if isinstance(stamp, datetime) else None,
            )
        )
    return rows


class VehicleStatusFileProvider(TelematicsProvider):
    """The daily feed, as a file the depot drops somewhere we can read.

    Stands in for the telematics API until there is one to call. Everything
    downstream — the readings, the history, the dockings they trigger — is the
    same either way, so swapping this for a real client changes one class.
    """

    def __init__(self, source: str) -> None:
        super().__init__(source)
        self.path = Path(source[len(FILE_SOURCE_PREFIX) :])

    @property
    def is_configured(self) -> bool:
        return self.path.exists()

    async def fetch(
        self, registration_nos: list[str]
    ) -> dict[str, tuple[int, datetime]]:
        if not self.is_configured:
            return {}
        wanted = {r.upper() for r in registration_nos}
        now = datetime.now(UTC)

        readings: dict[str, tuple[int, datetime]] = {}
        for row in read_vehicle_status(self.path):
            if row.odometer_km is None or row.registration_no not in wanted:
                continue
            stamp = row.recorded_at or now
            if stamp.tzinfo is None:
                stamp = stamp.replace(tzinfo=UTC)
            readings[row.registration_no] = (row.odometer_km, stamp)
        return readings
