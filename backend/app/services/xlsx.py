"""Excel exports.

The DMR month grid, and the two control charts whose blocks carry more text
than a 34px cell or a CSV suffix can show honestly — a depot reading "BD" on
screen still needs the actual breakdown description somewhere.
"""
from __future__ import annotations

from datetime import date as date_t
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.services import control_charts, dmr


def _bytes(workbook: Workbook) -> bytes:
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def dmr_month(
    *,
    site_code: str,
    month: str,
    dates: list[date_t],
    values: dict[str, list[float | None]],
) -> bytes:
    """Parameters down the rows, one column per day, Total last."""
    workbook = Workbook()
    sheet: Worksheet = workbook.active
    sheet.title = "DMR"

    sheet.append(["Daily Maintenance Report (DMR)"])
    sheet.append([f"Location :- {site_code}", None, f"Month :- {month}"])
    sheet.append([])

    header = ["Sl.No", "Parameters", *[d.isoformat() for d in dates], "Total"]
    sheet.append(header)
    header_row = sheet.max_row
    for cell in sheet[header_row]:
        cell.font = Font(bold=True)

    for parameter in dmr.PARAMETERS:
        row_values = values.get(parameter.key, [])
        total = dmr.monthly_total(parameter, row_values)
        sheet.append(
            [
                parameter.number,
                parameter.label,
                *[
                    None
                    if v is None
                    else (round(float(v), 1) if parameter.decimal else int(v))
                    for v in row_values
                ],
                None
                if total is None
                else (round(total, 1) if parameter.decimal else int(total)),
            ]
        )
        sheet.cell(row=sheet.max_row, column=len(header)).font = Font(bold=True)

    sheet.column_dimensions["A"].width = 6
    sheet.column_dimensions["B"].width = 42
    for col in range(3, len(header) + 1):
        sheet.column_dimensions[get_column_letter(col)].width = 7
    sheet.freeze_panes = sheet.cell(row=header_row + 1, column=3)
    for row in sheet.iter_rows(min_row=header_row, max_row=sheet.max_row, min_col=3):
        for cell in row:
            cell.alignment = Alignment(horizontal="center")

    return _bytes(workbook)


#: The red the breakdowns chart marks a block with — matches the PDF's
#: RED_FILL closely enough for the two to read as the same chart.
_BREAKDOWN_FILL = PatternFill(
    start_color="FDE2E1", end_color="FDE2E1", fill_type="solid"
)


def control_chart(chart: control_charts.Chart) -> bytes:
    """Bus down, days across, same as the screen — except a block holds its
    full text here instead of "BD2" with the rest only on a tooltip.

    Only meaningful for the charts that carry a title worth reading —
    driver complaints and breakdowns. Called for any chart, it still
    produces the grid, just with `value` standing in for the missing title.
    """
    workbook = Workbook()
    sheet: Worksheet = workbook.active
    sheet.title = chart.spec.title[:31] or "Chart"

    sheet.append([chart.spec.title])
    sheet.append([f"Location :- {chart.site_code}"])
    sheet.append(
        [f"{chart.from_date.isoformat()} to {chart.to_date.isoformat()}"]
    )
    sheet.append([chart.spec.legend])
    sheet.append([])

    header = ["Bus No", *[d.isoformat() for d in chart.dates]]
    sheet.append(header)
    header_row = sheet.max_row
    for cell in sheet[header_row]:
        cell.font = Font(bold=True)

    for row in chart.rows:
        sheet.append(
            [row.registration_no, *[c.title or c.value for c in row.cells]]
        )
        out_row = sheet.max_row
        sheet.cell(row=out_row, column=1).font = Font(bold=True)
        for i, c in enumerate(row.cells, start=2):
            if c.mark is control_charts.CellMark.breakdown:
                sheet.cell(row=out_row, column=i).fill = _BREAKDOWN_FILL

    sheet.column_dimensions["A"].width = 14
    for col in range(2, len(header) + 1):
        sheet.column_dimensions[get_column_letter(col)].width = 32
    sheet.freeze_panes = sheet.cell(row=header_row + 1, column=2)
    for row in sheet.iter_rows(min_row=header_row + 1, max_row=sheet.max_row):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    return _bytes(workbook)
