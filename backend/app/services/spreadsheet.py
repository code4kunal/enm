"""Reading uploaded spreadsheets.

Parsing belongs on the server so there is exactly one implementation and one
row-error report. The client uploads bytes and renders what comes back.
"""
from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass, field
from datetime import date as date_t
from datetime import datetime
from datetime import time as time_t
from decimal import Decimal

from app.errors import ValidationError

CSV_SUFFIXES = (".csv", ".txt")
XLSX_SUFFIXES = (".xlsx", ".xlsm")
XLS_SUFFIXES = (".xls",)


@dataclass(slots=True)
class SourceRow:
    """One data row, keyed by resolved column header.

    `number` is the row the user sees in Excel — blank rows are skipped for
    content but must not shift the numbering.
    """

    number: int
    values: dict[str, str]

    def get(self, column: str) -> str:
        return self.values.get(column, "")


@dataclass(slots=True)
class Sheet:
    name: str
    columns: list[str]
    rows: list[SourceRow]
    sheet_names: list[str] = field(default_factory=list)


def _cell_to_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, datetime):
        # A date-only cell arrives as midnight; keep it date-only so the ISO
        # parser downstream does not have to strip a spurious time.
        if value.time() == time_t(0, 0):
            return value.date().isoformat()
        return value.isoformat(timespec="seconds")
    if isinstance(value, date_t):
        return value.isoformat()
    if isinstance(value, time_t):
        return value.strftime("%H:%M")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, Decimal):
        return format(value.normalize(), "f")
    return str(value).strip()


def resolve_headers(raw: list[str]) -> list[str]:
    """Make every header addressable.

    Blanks become "Column 3" (1-based, as in the sheet) and duplicates get a
    "(2)" suffix, so a mapping can always name exactly one column.

    Internal whitespace is collapsed: real sheets wrap long headings across
    lines ("TYPE OF \nWORK"), and a saved mapping must not depend on where the
    author happened to press Alt+Enter.
    """
    out: list[str] = []
    seen: dict[str, int] = {}
    for index, value in enumerate(raw):
        name = re.sub(r"\s+", " ", value or "").strip()
        if not name:
            name = f"Column {index + 1}"
        count = seen.get(name.lower(), 0) + 1
        seen[name.lower()] = count
        out.append(name if count == 1 else f"{name} ({count})")
    return out


def _trim_trailing_blanks(grid: list[list[str]]) -> list[list[str]]:
    while grid and not any(cell for cell in grid[-1]):
        grid.pop()
    return grid


def _read_csv(data: bytes) -> tuple[list[str], list[list[str]]]:
    text = data.decode("utf-8-sig", errors="replace")
    try:
        dialect = csv.Sniffer().sniff(text[:4096], delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    grid = [
        [(cell or "").strip() for cell in row]
        for row in csv.reader(io.StringIO(text), dialect)
    ]
    return ["Sheet1"], grid


def _read_xlsx(data: bytes, sheet_name: str | None) -> tuple[list[str], list[list[str]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependency is pinned
        raise ValidationError(
            "This server cannot read .xlsx files (openpyxl is not installed)",
            {"file": "unsupported format"},
        ) from exc

    workbook = load_workbook(
        io.BytesIO(data), read_only=True, data_only=True, keep_links=False
    )
    try:
        names = list(workbook.sheetnames)
        if sheet_name and sheet_name not in names:
            raise ValidationError(
                f'The file has no sheet named "{sheet_name}"',
                {"sheet_name": "not found"},
            )
        worksheet = workbook[sheet_name] if sheet_name else workbook[names[0]]
        grid = [
            [_cell_to_text(cell) for cell in row]
            for row in worksheet.iter_rows(values_only=True)
        ]
        return names, grid
    finally:
        workbook.close()


def _read_xls(data: bytes, sheet_name: str | None) -> tuple[list[str], list[list[str]]]:
    try:
        import xlrd
    except ImportError as exc:  # pragma: no cover - dependency is pinned
        raise ValidationError(
            "This server cannot read legacy .xls files (xlrd is not installed). "
            "Save the sheet as .xlsx or .csv and upload again.",
            {"file": "unsupported format"},
        ) from exc

    book = xlrd.open_workbook(file_contents=data)
    names = book.sheet_names()
    if sheet_name and sheet_name not in names:
        raise ValidationError(
            f'The file has no sheet named "{sheet_name}"', {"sheet_name": "not found"}
        )
    sheet = book.sheet_by_name(sheet_name) if sheet_name else book.sheet_by_index(0)

    grid: list[list[str]] = []
    for r in range(sheet.nrows):
        row: list[str] = []
        for c in range(sheet.ncols):
            cell = sheet.cell(r, c)
            if cell.ctype == xlrd.XL_CELL_DATE:
                parts = xlrd.xldate_as_tuple(cell.value, book.datemode)
                row.append(_cell_to_text(datetime(*parts)))
            else:
                row.append(_cell_to_text(cell.value))
        grid.append(row)
    return names, grid


def read_sheet(
    *,
    file_name: str,
    data: bytes,
    sheet_name: str | None = None,
    header_row: int = 1,
    skip_rows: int = 0,
) -> Sheet:
    """Open an uploaded file and return its headers and data rows."""
    if not data:
        raise ValidationError("The uploaded file is empty", {"file": "empty"})

    lower = file_name.lower()
    if lower.endswith(CSV_SUFFIXES):
        names, grid = _read_csv(data)
    elif lower.endswith(XLSX_SUFFIXES):
        names, grid = _read_xlsx(data, sheet_name)
    elif lower.endswith(XLS_SUFFIXES):
        names, grid = _read_xls(data, sheet_name)
    else:
        raise ValidationError(
            f"{file_name} is not a spreadsheet — upload .xlsx, .xls or .csv",
            {"file": "unsupported format"},
        )

    grid = _trim_trailing_blanks(grid)
    if header_row > len(grid):
        raise ValidationError(
            f"Row {header_row} is past the end of the sheet",
            {"header_row": "past the end of the sheet"},
        )

    columns = resolve_headers(grid[header_row - 1])
    rows: list[SourceRow] = []
    # Row numbers are absolute sheet rows, so a blank row in the middle does not
    # shift what the user is told to look at in Excel.
    for offset, raw in enumerate(grid[header_row + skip_rows :]):
        number = header_row + skip_rows + offset + 1
        if not any(cell for cell in raw):
            continue
        values = {
            column: (raw[i] if i < len(raw) else "")
            for i, column in enumerate(columns)
        }
        rows.append(SourceRow(number=number, values=values))

    return Sheet(
        name=sheet_name or (names[0] if names else "Sheet1"),
        columns=columns,
        rows=rows,
        sheet_names=names,
    )
